#!/usr/bin/env python3
"""Generate the synthetic XLSX fixture corpus for :core:parser tests.

These files are the ``synthetic`` producer in the fixture taxonomy
(see core/parser/src/test/resources/fixtures/FIXTURES.md). Each one targets a
specific parser concern (values, shared strings, styles, merges, frozen panes,
dates, multiple sheets, sparsity) with **known golden values** that the T3-T7
golden tests assert against.

Note: openpyxl 3.1.5 ALWAYS writes string cells as ``inlineStr`` and never emits
``xl/sharedStrings.xml`` (see cell/_writer.py). So ``strings-shared.xlsx`` — the
one fixture whose whole purpose is to exercise the shared-string table (T4) — is
hand-crafted as minimal OOXML here instead of via openpyxl. See FIXTURES.md.

Setup:
    pip install openpyxl        # tested with openpyxl 3.1.5

Usage:
    python3 tools/gen_fixtures.py          # the 8 small golden fixtures
    python3 tools/gen_fixtures.py big       # big-50k-rows.xlsx (M2 perf target)

The 50k-row perf fixture is producer-agnostic (scale, not producer variance) and
is generated separately because it is large and slow.

Re-running is safe and (near-)deterministic: document timestamps are pinned so
regenerated files do not churn in git. The exact ZIP bytes may still differ
across openpyxl versions, but the parsed values the tests rely on will not.
"""

from __future__ import annotations

import datetime as dt
import sys
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# fixtures/synthetic/, resolved relative to this script so it runs from anywhere.
REPO_ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = REPO_ROOT / "core/parser/src/test/resources/fixtures/synthetic"

# Pinned document metadata -> reproducible output (no timestamp churn).
_PINNED = dt.datetime(2024, 1, 1, 0, 0, 0)


def _new_workbook(first_sheet: str = "Sheet1") -> Workbook:
    """Create a workbook with a single named sheet and pinned properties."""
    wb = Workbook()
    wb.active.title = first_sheet
    wb.properties.creator = "Darcha fixture generator"
    wb.properties.created = _PINNED
    wb.properties.modified = _PINNED
    return wb


def _save(wb: Workbook, name: str) -> None:
    path = OUT_DIR / name
    wb.save(path)
    print(f"  wrote {path.relative_to(REPO_ROOT)}")


# Fixed ZIP entry timestamp so hand-crafted files are byte-reproducible.
_ZIP_DATE = (2024, 1, 1, 0, 0, 0)


def _write_ooxml(name: str, parts: list[tuple[str, str]], note: str = "") -> None:
    """Write a raw OOXML package (an .xlsx ZIP) from ordered (path, xml) parts.

    Used only where openpyxl cannot produce the XML we need to test (shared
    strings). Entry timestamps are pinned for reproducibility.
    """
    path = OUT_DIR / name
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for part_name, content in parts:
            info = zipfile.ZipInfo(part_name, date_time=_ZIP_DATE)
            info.compress_type = zipfile.ZIP_DEFLATED
            zf.writestr(info, content.lstrip("\n"))
    print(f"  wrote {path.relative_to(REPO_ROOT)}  {note}".rstrip())


def gen_values_basic() -> None:
    """Numbers, text and booleans in a small dense grid."""
    wb = _new_workbook()
    ws = wb.active
    ws["A1"], ws["B1"], ws["C1"] = "Name", "Age", "Active"
    ws["A2"], ws["B2"], ws["C2"] = "Alice", 30, True
    ws["A3"], ws["B3"], ws["C3"] = "Bob", 25.5, False
    ws["A4"], ws["B4"], ws["C4"] = "Carol", 0, True
    _save(wb, "values-basic.xlsx")


def gen_strings_shared() -> None:
    """Hand-crafted minimal OOXML with a REAL shared string table (T4).

    openpyxl cannot emit sharedStrings.xml (it always writes inlineStr), so this
    one fixture is built as raw OOXML. Column A rows 1-7 reference the shared
    table by index:

        table:  0=fruit  1=apple  2=banana  3=cherry   (uniqueCount=4)
        A1..A7: 0, 1, 2, 1, 3, 2, 1                     (count=7)

    So A2, A4, A7 all resolve to "apple" via the same index 1.
    """
    content_types = """
<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>
  <Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>
  <Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>
  <Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>
</Types>
"""
    root_rels = """
<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>
  <Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>
</Relationships>
"""
    core = """
<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
  <dc:creator>Darcha fixture generator</dc:creator>
  <dcterms:created xsi:type="dcterms:W3CDTF">2024-01-01T00:00:00Z</dcterms:created>
  <dcterms:modified xsi:type="dcterms:W3CDTF">2024-01-01T00:00:00Z</dcterms:modified>
</cp:coreProperties>
"""
    app = """
<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties">
  <Application>Darcha fixture generator</Application>
</Properties>
"""
    workbook = """
<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets>
    <sheet name="Sheet1" sheetId="1" r:id="rId1"/>
  </sheets>
</workbook>
"""
    workbook_rels = """
<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
  <Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/>
</Relationships>
"""
    styles = """
<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>
  <fills count="2"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill></fills>
  <borders count="1"><border/></borders>
  <cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
  <cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>
  <cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>
</styleSheet>
"""
    shared_strings = """
<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="7" uniqueCount="4">
  <si><t>fruit</t></si>
  <si><t>apple</t></si>
  <si><t>banana</t></si>
  <si><t>cherry</t></si>
</sst>
"""
    indices = [0, 1, 2, 1, 3, 2, 1]
    rows = "".join(
        f'<row r="{i}"><c r="A{i}" t="s"><v>{idx}</v></c></row>'
        for i, idx in enumerate(indices, start=1)
    )
    sheet1 = f"""
<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <dimension ref="A1:A7"/>
  <sheetData>{rows}</sheetData>
</worksheet>
"""

    _write_ooxml(
        "strings-shared.xlsx",
        [
            ("[Content_Types].xml", content_types),
            ("_rels/.rels", root_rels),
            ("docProps/core.xml", core),
            ("docProps/app.xml", app),
            ("xl/workbook.xml", workbook),
            ("xl/_rels/workbook.xml.rels", workbook_rels),
            ("xl/styles.xml", styles),
            ("xl/sharedStrings.xml", shared_strings),
            ("xl/worksheets/sheet1.xml", sheet1),
        ],
        note="(hand-crafted shared strings)",
    )


def gen_styles_basic() -> None:
    """Bold, italic, font color, solid fill and alignments."""
    wb = _new_workbook()
    ws = wb.active

    ws["A1"] = "Bold"
    ws["A1"].font = Font(bold=True)

    ws["A2"] = "Italic"
    ws["A2"].font = Font(italic=True)

    ws["A3"] = "Red"
    ws["A3"].font = Font(color="FFFF0000")  # ARGB: opaque red

    ws["B1"] = "Fill"
    ws["B1"].fill = PatternFill(
        start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"  # yellow
    )

    ws["C1"] = "Center"
    ws["C1"].alignment = Alignment(horizontal="center", vertical="center")

    ws["C2"] = "Right"
    ws["C2"].alignment = Alignment(horizontal="right")

    _save(wb, "styles-basic.xlsx")


def gen_merged() -> None:
    """Three merged ranges: a horizontal title, a vertical side, a block."""
    wb = _new_workbook()
    ws = wb.active
    ws["A1"] = "Title"
    ws.merge_cells("A1:C1")
    ws["A2"] = "Side"
    ws.merge_cells("A2:A4")
    ws["B2"] = "Block"
    ws.merge_cells("B2:C3")
    _save(wb, "merged.xlsx")


def gen_frozen() -> None:
    """Freeze the first row and first column (pane anchor B2)."""
    wb = _new_workbook()
    ws = wb.active
    ws["A1"], ws["B1"], ws["C1"] = "Corner", "H1", "H2"
    ws["A2"], ws["B2"], ws["C2"] = "R1", 1, 2
    ws["A3"], ws["B3"], ws["C3"] = "R2", 3, 4
    ws.freeze_panes = "B2"  # xSplit=1, ySplit=1
    _save(wb, "frozen.xlsx")


def gen_dates() -> None:
    """Date/time/datetime cells across builtin and custom number formats.

    A1 builtin 14 (date), A2 builtin 21 (time), A3 builtin 22 (datetime),
    A4 custom 'yyyy-mm-dd' (date detected via y/m/d tokens).
    """
    wb = _new_workbook()
    ws = wb.active

    ws["A1"] = dt.date(2024, 1, 15)
    ws["A1"].number_format = "mm-dd-yy"  # builtin id 14

    ws["A2"] = dt.time(13, 30, 0)
    ws["A2"].number_format = "h:mm:ss"  # builtin id 21

    ws["A3"] = dt.datetime(2024, 1, 15, 13, 30, 0)
    ws["A3"].number_format = "m/d/yy h:mm"  # builtin id 22

    ws["A4"] = dt.date(2024, 12, 31)
    ws["A4"].number_format = "yyyy-mm-dd"  # custom (>=164), date by tokens

    _save(wb, "dates.xlsx")


def gen_multisheet() -> None:
    """Three sheets in order, including a non-ASCII (Cyrillic) name."""
    wb = _new_workbook(first_sheet="Jadval 1")
    wb.active["A1"] = "birinchi"

    ws2 = wb.create_sheet("Narxlar")
    ws2["A1"] = "ikkinchi"

    ws3 = wb.create_sheet("Ҳисобот")  # non-ASCII UTF-8 sheet name
    ws3["A1"] = "uchinchi"

    _save(wb, "multisheet.xlsx")


def gen_sparse_gaps() -> None:
    """Only three populated cells with large gaps: A1, C5, AA100.

    Column AA is index 26 (0-based); row 100 is index 99. The <dimension>
    element will claim A1:AA100, but only 3 <c> elements exist.
    """
    wb = _new_workbook()
    ws = wb.active
    ws["A1"] = "start"
    ws["C5"] = 42
    ws["AA100"] = "end"
    _save(wb, "sparse-gaps.xlsx")


def gen_text_contrast() -> None:
    """The six font-colour cases a viewer with its own themes has to tell apart (T28).

    Every row is a different answer to one question: did the document choose this
    colour, and did it also choose what sits behind it?

        A1  theme=1 text, no fill      -> chose NOTHING. sysClr windowText.
                                         The common case: every real Excel file.
        A2  explicit black, no fill    -> chose black, against a background it
                                         does not control. Invisible in dark.
        A3  explicit white, no fill    -> the inverse, invisible in LIGHT mode.
                                         Broken since v1.0 and never noticed,
                                         because nothing in the corpus had it.
        A4  black on yellow fill       -> chose BOTH. Must render as written.
        A5  white on dark-blue fill    -> chose both, the other way round.
        A6  grey #999999, no fill      -> chose a quiet colour on purpose.
                                         Must NOT be "rescued" into full contrast.
        A7  theme text on YELLOW fill  -> chose the background but not the text.
                                         Needs DARK text in both app themes.
        A8  theme text on NAVY fill    -> the same, the other polarity.
                                         Needs LIGHT text in both app themes.

    A1 is what T28 changed in the parser (theme 1 -> null, "no choice"). A2/A3
    are what the renderer's contrast rule catches. A4/A5/A6 are the cases that
    prove neither mechanism overreaches -- they are the reason this fixture has
    six rows instead of two.
    """
    from openpyxl.styles import Color as XlColor

    wb = _new_workbook()
    ws = wb.active
    ws["A1"] = "theme text"
    ws["A1"].font = Font(color=XlColor(theme=1, tint=0.0))
    ws["A2"] = "explicit black"
    ws["A2"].font = Font(color="FF000000")
    ws["A3"] = "explicit white"
    ws["A3"].font = Font(color="FFFFFFFF")
    ws["A4"] = "black on yellow"
    ws["A4"].font = Font(color="FF000000")
    ws["A4"].fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    ws["A5"] = "white on navy"
    ws["A5"].font = Font(color="FFFFFFFF")
    ws["A5"].fill = PatternFill(start_color="FF1F3864", end_color="FF1F3864", fill_type="solid")
    ws["A6"] = "quiet grey"
    ws["A6"].font = Font(color="FF999999")

    # A7/A8 -- the combination the corpus only had by accident, in
    # synthetic/styles-basic.xlsx B1, until it shipped a regression (v1.1.0).
    #
    # The document declines to choose a font colour (theme 1 = the system's text
    # colour) but DOES choose a fill. So the background is the author's while the
    # foreground is ours, and the readable colour depends entirely on the fill:
    # dark text on the yellow, light text on the navy, in EITHER app theme. Both
    # polarities are here because getting one right by luck is exactly how this
    # was missed the first time.
    ws["A7"] = "theme on light fill"
    ws["A7"].font = Font(color=XlColor(theme=1, tint=0.0))
    ws["A7"].fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    ws["A8"] = "theme on dark fill"
    ws["A8"].font = Font(color=XlColor(theme=1, tint=0.0))
    ws["A8"].fill = PatternFill(start_color="FF1F3864", end_color="FF1F3864", fill_type="solid")

    ws.column_dimensions["A"].width = 22
    _save(wb, "text-contrast.xlsx")


def gen_ods_renamed() -> None:
    """A real OpenDocument spreadsheet carrying an .xlsx name (T27).

    Darcha is offered for any file whose path ends in .xlsx (see the manifest's
    pathPattern filters), so a renamed .ods genuinely reaches the app. It is an
    intact spreadsheet of the wrong kind -- ErrorKind.Unsupported, not Corrupted.

    Hand-built rather than exported, because no LibreOffice is available on the
    machine that generated the corpus. It follows OpenDocument v1.2 section 3.3,
    which is what ContainerDetector relies on:

        * ``mimetype`` is the FIRST entry in the archive
        * it is STORED (compress_type=0), never deflated
        * it carries NO extra field

    Those three rules put the media type at a fixed byte offset, which is why
    detection costs one short read instead of opening the archive. The layout is
    asserted byte-by-byte in ContainerDetectorTest rather than taken on trust --
    a fixture built by the same author as the detector proves nothing otherwise.

    Replace with genuine LibreOffice or Google Sheets output when T30 fills the
    producer folders; the golden values should not change if this is right.
    """
    mimetype = b"application/vnd.oasis.opendocument.spreadsheet"
    manifest = """<?xml version="1.0" encoding="UTF-8"?>
<manifest:manifest xmlns:manifest="urn:oasis:names:tc:opendocument:xmlns:manifest:1.0" manifest:version="1.2">
  <manifest:file-entry manifest:full-path="/" manifest:version="1.2" manifest:media-type="application/vnd.oasis.opendocument.spreadsheet"/>
  <manifest:file-entry manifest:full-path="content.xml" manifest:media-type="text/xml"/>
  <manifest:file-entry manifest:full-path="styles.xml" manifest:media-type="text/xml"/>
</manifest:manifest>
"""
    content = """<?xml version="1.0" encoding="UTF-8"?>
<office:document-content xmlns:office="urn:oasis:names:tc:opendocument:xmlns:office:1.0" xmlns:table="urn:oasis:names:tc:opendocument:xmlns:table:1.0" xmlns:text="urn:oasis:names:tc:opendocument:xmlns:text:1.0" office:version="1.2">
  <office:body>
    <office:spreadsheet>
      <table:table table:name="Sheet1">
        <table:table-row>
          <table:table-cell office:value-type="string"><text:p>Nomi</text:p></table:table-cell>
          <table:table-cell office:value-type="string"><text:p>Narxi</text:p></table:table-cell>
        </table:table-row>
        <table:table-row>
          <table:table-cell office:value-type="string"><text:p>Olma</text:p></table:table-cell>
          <table:table-cell office:value-type="float" office:value="5000"><text:p>5000</text:p></table:table-cell>
        </table:table-row>
      </table:table>
    </office:spreadsheet>
  </office:body>
</office:document-content>
"""
    styles = """<?xml version="1.0" encoding="UTF-8"?>
<office:document-styles xmlns:office="urn:oasis:names:tc:opendocument:xmlns:office:1.0" office:version="1.2"/>
"""

    path = OUT_DIR / "ods-renamed.xlsx"
    with zipfile.ZipFile(path, "w") as zf:
        # The mimetype entry FIRST and STORED, with no extra field. Written via an
        # explicit ZipInfo so none of that is left to zipfile's defaults.
        info = zipfile.ZipInfo("mimetype", date_time=_ZIP_DATE)
        info.compress_type = zipfile.ZIP_STORED
        info.extra = b""
        zf.writestr(info, mimetype)

        for part_name, text in (
            ("META-INF/manifest.xml", manifest),
            ("content.xml", content),
            ("styles.xml", styles),
        ):
            part = zipfile.ZipInfo(part_name, date_time=_ZIP_DATE)
            part.compress_type = zipfile.ZIP_DEFLATED
            zf.writestr(part, text)

    _assert_odf_layout(path, mimetype)
    print(f"  wrote {path.relative_to(REPO_ROOT)}  (ODF package, .xlsx name)")


def _assert_odf_layout(path, mimetype: bytes) -> None:
    """Fail loudly unless the written file really has the ODF layout.

    zipfile is free to add extra fields or pick a compression method; if it ever
    does, the fixture would silently stop testing what it exists to test. Checked
    here at generation time as well as in the Kotlin test, because a fixture that
    quietly drifts is worse than no fixture.
    """
    head = path.read_bytes()[:128]
    method = int.from_bytes(head[8:10], "little")
    name_len = int.from_bytes(head[26:28], "little")
    extra_len = int.from_bytes(head[28:30], "little")
    declared = int.from_bytes(head[18:22], "little")
    name = head[30:30 + name_len]
    body = head[30 + name_len + extra_len:][:declared]

    problems = []
    if head[:4] != b"PK\x03\x04":
        problems.append("not a ZIP local file header")
    if method != 0:
        problems.append(f"mimetype is compressed (method={method}), must be STORED")
    if extra_len != 0:
        problems.append(f"mimetype has a {extra_len}-byte extra field, must have none")
    if name != b"mimetype":
        problems.append(f"first entry is {name!r}, must be b'mimetype'")
    if body != mimetype:
        problems.append(f"media type at the fixed offset is {body!r}")
    if problems:
        raise SystemExit("ODF layout is wrong: " + "; ".join(problems))


def gen_column_widths() -> None:
    """Custom column widths and row heights — the layout path (T15.6).

    Almost every real business spreadsheet sets column widths, so this is the
    normal case rather than an edge case, and until this fixture existed nothing
    in the corpus exercised <cols> at any size. Ten rows, so a small chunkSize
    spans several chunks and the tests can prove every chunk carries the column
    layout while row heights arrive with their own rows.

    Widths (character units): A, C and D custom; B left at the sheet default.
    Heights (points): rows 1 and 7 custom; the rest default.
    """
    wb = _new_workbook()
    ws = wb.active

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["C"].width = 4.5
    ws.column_dimensions["D"].width = 18
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[7].height = 28

    ws.append(["name", "qty", "id", "note"])
    for i in range(1, 10):
        ws.append([f"row {i}", i * 10, i, "-"])

    _save(wb, "column-widths.xlsx")


def gen_big_50k() -> None:
    """A large sheet (~50k rows) for the M2 performance target.

    Producer-agnostic: this is a scale/perf fixture, not a producer-variance one,
    so openpyxl is fine. Written in write-only mode to stay memory-cheap. Mostly
    numeric to keep the file size modest; one repeated text column and a boolean
    exercise the string/bool paths at scale.
    """
    from openpyxl import Workbook as WriteOnlyWorkbook

    ROWS = 50_000
    CATEGORIES = ("alpha", "beta", "gamma", "delta", "epsilon", "zeta")

    wb = WriteOnlyWorkbook(write_only=True)
    ws = wb.create_sheet("Data")
    ws.append(["id", "value", "delta", "count", "pct", "category", "flag"])
    for i in range(1, ROWS + 1):
        ws.append([
            i,
            (i * 7) % 1000,
            round((i % 97) / 97.0, 4),
            i * 3,
            round((i % 100) / 100.0, 2),
            CATEGORIES[i % len(CATEGORIES)],
            i % 2 == 0,
        ])
    wb.properties.creator = "Darcha fixture generator"
    wb.properties.created = _PINNED
    wb.properties.modified = _PINNED

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUT_DIR / "big-50k-rows.xlsx"
    wb.save(path)
    print(f"  wrote {path.relative_to(REPO_ROOT)}  ({ROWS} rows, perf fixture)")


def gen_big_50k_wide() -> None:
    """`big-50k-rows.xlsx` plus custom column widths — the T15.6 device check.

    Verifying "a partial paint does not reflow" needs a file large enough that
    partial paints actually happen AND wide enough that a reflow would be
    obvious. The small `column-widths.xlsx` covers the golden values; this one
    covers the eye.

    Deliberately **not** committed: it is a measurement aid, not a golden
    fixture, and a second 1.8 MB file would double the corpus for no assertion.
    Regenerate it when re-measuring, then delete it. See docs/PERF.md.
    """
    from openpyxl import Workbook as WriteOnlyWorkbook
    from openpyxl.utils import get_column_letter

    ROWS = 50_000
    CATEGORIES = ("alpha", "beta", "gamma", "delta", "epsilon", "zeta")
    # Wildly non-default, so an unstyled first paint would be unmistakable.
    WIDTHS = (42.0, 4.0, 4.0, 30.0, 4.0, 26.0, 4.0)

    wb = WriteOnlyWorkbook(write_only=True)
    ws = wb.create_sheet("Data")
    for i, width in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.append(["id", "value", "delta", "count", "pct", "category", "flag"])
    for i in range(1, ROWS + 1):
        ws.append([
            i,
            (i * 7) % 1000,
            round((i % 97) / 97.0, 4),
            i * 3,
            round((i % 100) / 100.0, 2),
            CATEGORIES[i % len(CATEGORIES)],
            i % 2 == 0,
        ])
    wb.properties.creator = "Darcha fixture generator"
    wb.properties.created = _PINNED
    wb.properties.modified = _PINNED

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUT_DIR / "big-50k-wide.xlsx"
    wb.save(path)
    print(f"  wrote {path.relative_to(REPO_ROOT)}  ({ROWS} rows, custom widths, NOT committed)")


def gen_styled_20k() -> None:
    """A heavily styled 20k-row sheet — the T17 render/cache measurement.

    Every row carries a fill, a font colour, bold/italic and an alignment,
    cycling through eight combinations, plus a repeated category column so the
    *same text* appears under many styles. That is the case that multiplies text
    cache keys once the key includes the style id.

    Deliberately **not** committed, like `big-50k-wide.xlsx`: it is a measurement
    aid, not a golden fixture. Regenerate when re-measuring, then delete it.
    See docs/PERF.md.
    """
    from openpyxl.styles import Alignment, Font, PatternFill

    ROWS = 20_000
    CATEGORIES = ("alpha", "beta", "gamma", "delta", "epsilon", "zeta")
    FILLS = ("FFFFF2CC", "FFD9E1F2", "FFE2EFDA", "FFFCE4D6")

    wb = _new_workbook(first_sheet="Styled")
    ws = wb.active
    ws.append(["id", "value", "delta", "count", "pct", "category", "when"])
    for i in range(1, ROWS + 1):
        ws.append([
            i,
            (i * 7) % 1000,
            round((i % 97) / 97.0, 4),
            i * 3,
            round((i % 100) / 100.0, 2),
            CATEGORIES[i % len(CATEGORIES)],
            dt.date(2024, 1, 1) + dt.timedelta(days=i % 365),
        ])
        k = i % 8
        for cell in ws[i + 1]:
            cell.font = Font(
                bold=(k % 2 == 0),
                italic=(k % 3 == 0),
                color=("FFC00000" if k % 4 == 0 else "FF1F3864"),
            )
            cell.fill = PatternFill("solid", fgColor=FILLS[k % 4])
            cell.alignment = Alignment(horizontal=("center" if k % 3 == 0 else "right"))
        ws.cell(row=i + 1, column=7).number_format = "yyyy-mm-dd"

    _save(wb, "styled-20k.xlsx")
    print("    (20000 rows, 8 style combinations, NOT committed)")


def gen_big_merged() -> None:
    """A 50k-row sheet with merges — the T18 progressive-paint check.

    `<mergeCells>` follows `<sheetData>`, so merges cannot reach a partial paint
    (see TECH_SPEC §7). This file makes that visible: a merged banner at the top
    and a merged section header every 100 rows, on a sheet big enough that the
    grid is on screen for seconds before the merges arrive.

    Deliberately **not** committed, like the other measurement aids.
    """
    from openpyxl.styles import Alignment, Font, PatternFill

    ROWS = 50_000
    COLUMNS = 7
    SECTION_EVERY = 100

    wb = _new_workbook(first_sheet="Report")
    ws = wb.active
    ws["A1"] = "Yillik hisobot — 2024"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = PatternFill("solid", fgColor="FFD9E1F2")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=COLUMNS)

    ws.append(["id", "value", "delta", "count", "pct", "category", "note"])
    row = 2
    for i in range(1, ROWS + 1):
        row += 1
        if i % SECTION_EVERY == 0:
            ws.cell(row=row, column=1, value=f"Bo'lim {i // SECTION_EVERY}")
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="FFE2EFDA")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLUMNS)
            continue
        for c, v in enumerate(
            [i, (i * 7) % 1000, round((i % 97) / 97.0, 4), i * 3,
             round((i % 100) / 100.0, 2), f"kat-{i % 6}", "-"],
            start=1,
        ):
            ws.cell(row=row, column=c, value=v)

    _save(wb, "big-merged.xlsx")
    print(f"    ({ROWS} rows, {1 + ROWS // SECTION_EVERY} merges, NOT committed)")


def gen_big_5mb() -> None:
    """A ~4.7 MB sheet — the upper end of the §5 "typical files (< 5 MB)" target.

    That target was written for files up to 5 MB and, until this file existed,
    had only ever been measured up to 1.78 MB. This closes the untested half.

    **Wide, not deep — 40 columns x 11,500 rows.** Two reasons, and the first is
    the one that decided it. The parser emits the first ~200 rows immediately
    (TECH_SPEC §7), so on a 40-column sheet that first chunk carries **8,000
    cells** rather than the 1,400 a 7-column sheet would give it: at this size
    the interesting question is what the *first paint* costs, and a wide sheet
    asks it far harder. Deep-and-narrow would have needed roughly 64,000 rows of
    7 columns for the same bytes and tested the same path with a fifth of the
    work up front. Second, a 5 MB business spreadsheet is wide in real life.

    **Staying under the cap was a content problem, not a shape one.** Cells are
    `rows x cols` whichever way the sheet leans, so the only way to reach 5 MB
    with room to spare is to make each cell cost more bytes — which means text.
    `big-50k-rows.xlsx` is mostly numeric and runs about 5 bytes per cell, so
    5 MB of it would be ~950,000 cells, within a whisker of the 1,000,000 cap
    (§9.1) — a `TooLarge` test wearing a parse-speed costume. Every third column
    here is a multi-word description, which measures ~11 bytes per cell and puts
    4.7 MB at **460,000 cells, 46% of the cap**.

    Text is drawn from a fixed vocabulary with a pinned seed, so the file is
    byte-reproducible, and it is Uzbek because that is what this app's files
    look like — it also keeps multi-byte UTF-8 in the hot path.

    Deliberately **not** committed, like the other measurement aids: 4.7 MB
    would nearly quadruple the corpus for no assertion. Generate it, measure,
    delete. See docs/PERF.md.
    """
    import random

    from openpyxl import Workbook as WriteOnlyWorkbook

    ROWS = 11_500
    COLUMNS = 40
    DESCRIPTION_EVERY = 3
    WORDS = (
        "Toshkent", "Samarqand", "Buxoro", "Andijon", "Farg'ona", "Namangan",
        "Qarshi", "Nukus", "Xiva", "hisobot", "oylik", "kirim", "chiqim",
        "qoldiq", "jami", "filial", "ombor", "yetkazib berish", "shartnoma",
        "to'lov", "balans", "mijoz", "mahsulot", "narx", "miqdor", "sana",
        "izoh", "holat", "tasdiqlangan", "kutilmoqda",
    )

    rng = random.Random(20260805)
    wb = WriteOnlyWorkbook(write_only=True)
    ws = wb.create_sheet("Data")
    ws.append([f"ustun_{c}" for c in range(1, COLUMNS + 1)])
    for i in range(1, ROWS + 1):
        row = []
        for c in range(COLUMNS):
            if c % DESCRIPTION_EVERY == 0:
                words = " ".join(rng.choice(WORDS) for _ in range(rng.randint(6, 12)))
                row.append(f"{words} #{i}")
            elif c % 5 == 1:
                row.append(round(rng.random() * 10_000, 2))
            elif c % 5 == 2:
                row.append(rng.randint(1, 999_999))
            elif c % 5 == 3:
                row.append(f"{rng.choice(WORDS)}-{rng.randint(1000, 9999)}")
            else:
                row.append(rng.choice(WORDS))
        ws.append(row)
    wb.properties.creator = "Darcha fixture generator"
    wb.properties.created = _PINNED
    wb.properties.modified = _PINNED

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUT_DIR / "big-5mb-wide.xlsx"
    wb.save(path)
    cells = (ROWS + 1) * COLUMNS
    size = path.stat().st_size
    print(
        f"  wrote {path.relative_to(REPO_ROOT)}  "
        f"({size / 1048576:.2f} MB, {ROWS} rows x {COLUMNS} cols = {cells:,} cells, NOT committed)"
    )


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    if "big" in sys.argv[1:]:
        print("Generating the 50k-row performance fixtures ...")
        gen_big_50k()
        gen_big_50k_wide()
        gen_styled_20k()
        gen_big_merged()
        gen_big_5mb()
        print("Done.")
        return
    print(f"Generating synthetic fixtures into {OUT_DIR.relative_to(REPO_ROOT)} ...")
    gen_values_basic()
    gen_strings_shared()
    gen_styles_basic()
    gen_merged()
    gen_frozen()
    gen_dates()
    gen_multisheet()
    gen_sparse_gaps()
    gen_column_widths()
    gen_ods_renamed()
    print("Done: 10 fixtures generated.")


if __name__ == "__main__":
    main()
